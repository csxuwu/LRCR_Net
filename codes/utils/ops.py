

import os
import torch
from codes.utils import img_processing
from openpyxl import workbook
from openpyxl import load_workbook


def create_folder(path):
	if not os.path.exists(path):
		os.makedirs(path)
		print('creating :{}'.format(path))


def prepro_save(img):
	img = img.to(torch.device('cpu'))
	img = img.squeeze()
	img = img.mul_(255).add_(0.5).clamp_(0, 255).permute(1, 2, 0).type(torch.uint8)
	img = img.detach().numpy()
	return img

def create_excel(excel_name, excel_header):

	if os.path.exists(excel_name):
		excel = load_workbook(excel_name)
		excel_active = excel.active
	else:
		excel = workbook.Workbook()
		excel_active = excel.active
		# excel_active.append(
		# 	['global_steps', 'epoch', 'step', 'losses','psnr','ssim','lr'])
		excel_active.append(excel_header)
	print('Create excel:{}'.format(excel_name))
	return excel,excel_active


def numerical_normalizaiton(val):
	'''

	:param val:
	:return:
	'''
	val_str = str(val.item())
	b = val_str.split('.')
	l=len(b[0])

	out = val*pow(0.1,l-1)

	return out


def img_rename(img_path, out_path):

	no = 1
	imgs = os.listdir(img_path)
	for img in imgs:
		name = img.split('_')[0]
		img_name = name + '.jpg'
		print(no)
		os.rename(os.path.join(img_path,img),os.path.join(out_path,img_name))
		no += 1
		print('{} has renamed.'.format(img))
